import sys
import threading
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

import yfinance as yf
import pandas as pd
import numpy as np
from scipy import stats
import openpyxl

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QSpinBox, QDoubleSpinBox, QPushButton,
    QTableWidget, QTableWidgetItem, QDateEdit, QComboBox, QTextEdit
)
from PyQt6.QtCore import Qt, QDate, pyqtSignal, QThread
from PyQt6.QtGui import QFont

# =========================================================
# DATA CACHE
# =========================================================
class DataCache:
    def __init__(self):
        self.ticker_prices = {}  # {ticker: pd.Series}
        self.index_prices = None  # pd.Series
        self.last_beta_history = None
        self.last_vol_term = None
        self.last_tickers = None
        self.last_index = None
        self.pull_timestamp = None
    
    def needs_repull(self, tickers, index, beta_history, vol_term):
        """Check if data needs to be re-pulled"""
        # Always repull if tickers or index changed
        if set(tickers) != set(self.last_tickers or []) or index != self.last_index:
            return True
        
        # Repull if vol_term increased beyond years_available
        if self.last_vol_term and vol_term > self.last_vol_term:
            # Would need to check years_available, but for now assume repull if vol increases significantly
            if vol_term > 5.5:  # Beyond typical 5-year history
                return True
        
        return False
    
    def cache(self, ticker_prices_dict, index_prices_s, tickers, index, beta_history, vol_term):
        """Store data in cache"""
        self.ticker_prices = ticker_prices_dict
        self.index_prices = index_prices_s
        self.last_tickers = tickers
        self.last_index = index
        self.last_beta_history = beta_history
        self.last_vol_term = vol_term
        self.pull_timestamp = datetime.now()

# =========================================================
# WORKER THREAD FOR DATA PULLING
# =========================================================
class DataPullWorker(QThread):
    finished = pyqtSignal()
    error = pyqtSignal(str)
    results = pyqtSignal(list)  # List of dicts with results
    
    def __init__(self, tickers, index_ticker, valuation_date, beta_history, vol_term, cache):
        super().__init__()
        self.tickers = tickers
        self.index_ticker = index_ticker
        self.valuation_date = pd.Timestamp(valuation_date)  # Convert to Timestamp
        self.beta_history = beta_history
        self.vol_term = vol_term
        self.cache = cache
    
    def run(self):
        try:
            # Calculate date window
            end_date = self.valuation_date
            start_date = end_date - relativedelta(months=int(self.beta_history * 12)) - timedelta(days=30)
            
            # Fetch index data
            index_data = yf.download(self.index_ticker, start=start_date, end=end_date, progress=False)
            if len(index_data) == 0:
                self.error.emit(f"No data for index {self.index_ticker}")
                return
            
            index_prices = index_data[('Close', self.index_ticker)].sort_index(ascending=True)
            
            # Fetch ticker data
            ticker_prices_dict = {}
            for ticker in self.tickers:
                ticker_data = yf.download(ticker, start=start_date, end=end_date, progress=False)
                if len(ticker_data) > 0:
                    ticker_prices_dict[ticker] = ticker_data[('Close', ticker)].sort_index(ascending=True)
            
            # Cache the data
            self.cache.cache(ticker_prices_dict, index_prices, self.tickers, self.index_ticker, self.beta_history, self.vol_term)
            
            # Calculate betas and vols
            results = self._calculate_all(ticker_prices_dict, index_prices)
            self.results.emit(results)
            self.finished.emit()
        
        except Exception as e:
            self.error.emit(f"Data pull error: {str(e)}")
    
    def _calculate_all(self, ticker_prices_dict, index_prices):
        """Calculate betas and volatilities for all tickers"""
        results = []
        
        for ticker in self.tickers:
            if ticker not in ticker_prices_dict:
                results.append({
                    'Ticker': ticker,
                    '2yr Raw': None,
                    '2yr Adj': None,
                    '5yr Raw': None,
                    '5yr Adj': None,
                    'Volatility': None,
                    'YearsAvail': 0
                })
                continue
            
            ticker_prices = ticker_prices_dict[ticker]
            years_available = (ticker_prices.index[-1] - ticker_prices.index[0]).days / 365.25
            
            # Calculate betas
            beta_2yr_raw = self._calculate_beta(ticker_prices, index_prices, frequency='weekly')
            beta_2yr_adj = beta_2yr_raw * (2/3) + (1/3) if beta_2yr_raw else None
            
            beta_5yr_raw = self._calculate_beta(ticker_prices, index_prices, frequency='monthly')
            beta_5yr_adj = beta_5yr_raw * (2/3) + (1/3) if beta_5yr_raw else None
            
            # Calculate volatility
            volatility = self._calculate_volatility(ticker_prices, years_available)
            
            results.append({
                'Ticker': ticker,
                '2yr Raw': beta_2yr_raw,
                '2yr Adj': beta_2yr_adj,
                '5yr Raw': beta_5yr_raw,
                '5yr Adj': beta_5yr_adj,
                'Volatility': volatility,
                'YearsAvail': years_available
            })
        
        return results
    
    def _calculate_beta(self, ticker_prices, index_prices, frequency='weekly'):
        """Calculate beta (exact logic from beta_vol_loop15.py)"""
        if frequency == 'weekly':
            observations_needed = 104
            anchor = self.valuation_date - timedelta(days=(self.valuation_date.weekday() - 4) % 7)
        else:
            observations_needed = 60
            anchor = datetime(self.valuation_date.year, self.valuation_date.month, 1) - timedelta(days=1)
        
        sample_dates = []
        for i in range(observations_needed + 1):
            if frequency == 'weekly':
                target = anchor - timedelta(days=7 * i)
            else:
                month_offset = i
                year = anchor.year
                month = anchor.month - month_offset
                while month <= 0:
                    month += 12
                    year -= 1
                next_month = datetime(year, month, 1) + timedelta(days=32)
                target = datetime(next_month.year, next_month.month, 1) - timedelta(days=1)
            sample_dates.append(target)
        
        sample_dates.reverse()
        
        ticker_sample = []
        index_sample = []
        
        for target in sample_dates:
            t_date = self._find_closest_trading_date(target, pd.DatetimeIndex(ticker_prices.index))
            i_date = self._find_closest_trading_date(target, pd.DatetimeIndex(index_prices.index))
            
            if t_date is None or i_date is None:
                return None
            
            ticker_sample.append(ticker_prices[t_date])
            index_sample.append(index_prices[i_date])
        
        ticker_returns = np.diff(ticker_sample) / ticker_sample[:-1]
        index_returns = np.diff(index_sample) / index_sample[:-1]
        
        if len(ticker_returns) < 2:
            return None
        
        try:
            slope, _, _, _, _ = stats.linregress(index_returns, ticker_returns)
            return slope
        except:
            return None
    
    def _calculate_volatility(self, ticker_prices, years_available):
        """Calculate volatility (exact logic from beta_vol_loop15.py)"""
        vol_term_used = min(self.vol_term, years_available)
        
        if vol_term_used <= 0:
            return None
        
        n = len(ticker_prices)
        terms = []
        for d in ticker_prices.index:
            yearfrac = (self.valuation_date - pd.Timestamp(d)).days / 365.25
            term = round(yearfrac * 1000) / 1000
            terms.append(term)
        
        start_idx = 0
        for i in range(n):
            if terms[i] <= vol_term_used:
                start_idx = i
                break
        
        ret_count = n - start_idx
        if ret_count < 2:
            return None
        
        log_returns = np.log(ticker_prices.iloc[start_idx+1:].values / ticker_prices.iloc[start_idx:-1].values)
        vol = np.std(log_returns, ddof=0) * np.sqrt(252)
        
        return vol
    
    def _find_closest_trading_date(self, target_date, available_dates, direction='<='):
        """Find closest trading date to target"""
        if direction == '<=':
            valid = available_dates[available_dates <= target_date]
            if len(valid) > 0:
                return valid[-1]
        return None

# =========================================================
# MAIN WINDOW
# =========================================================
class BetaVolCalculator(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Project Canneberge — Beta & Vol Calculator")
        self.setGeometry(100, 100, 1200, 700)
        
        self.cache = DataCache()
        self.worker = None
        
        self.init_ui()
        self.load_index_universe()
        self.load_tickers_from_excel()
    
    def init_ui(self):
        """Initialize UI components"""
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        
        layout = QVBoxLayout()
        
        # Input panel
        input_layout = QHBoxLayout()
        
        # Valuation Date
        input_layout.addWidget(QLabel("Valuation Date:"))
        self.date_input = QDateEdit()
        self.date_input.setDate(QDate.currentDate())
        input_layout.addWidget(self.date_input)
        
        # Beta History
        input_layout.addWidget(QLabel("Beta History (years):"))
        self.beta_history_input = QDoubleSpinBox()
        self.beta_history_input.setValue(5.0)
        self.beta_history_input.setMinimum(1.0)
        self.beta_history_input.setMaximum(20.0)
        input_layout.addWidget(self.beta_history_input)
        
        # Volatility Term
        input_layout.addWidget(QLabel("Volatility Term (years):"))
        self.vol_term_input = QDoubleSpinBox()
        self.vol_term_input.setValue(3.0)
        self.vol_term_input.setMinimum(0.5)
        self.vol_term_input.setMaximum(20.0)
        input_layout.addWidget(self.vol_term_input)
        
        # Index
        input_layout.addWidget(QLabel("Index:"))
        self.index_combo = QComboBox()
        input_layout.addWidget(self.index_combo)
        
        # Calculate button
        self.calculate_btn = QPushButton("CALCULATE")
        self.calculate_btn.clicked.connect(self.on_calculate_clicked)
        input_layout.addWidget(self.calculate_btn)
        
        input_layout.addStretch()
        layout.addLayout(input_layout)
        
        # Tickers input
        ticker_layout = QHBoxLayout()
        ticker_layout.addWidget(QLabel("Tickers (comma-separated):"))
        self.tickers_input = QLineEdit()
        self.tickers_input.setText("RKLB, AMZN, FLY, ASTS, GOOG, IRDM, PLTR, SOUN, NBIS")
        ticker_layout.addWidget(self.tickers_input)
        layout.addLayout(ticker_layout)
        
        # Results table
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(8)
        self.results_table.setHorizontalHeaderLabels([
            "Ticker", "2yr Raw", "2yr Adj", "5yr Raw", "5yr Adj", "Volatility", "YearsAvail", "Status"
        ])
        layout.addWidget(self.results_table)
        
        # Status bar
        status_layout = QHBoxLayout()
        self.status_label = QLabel("Ready")
        status_layout.addWidget(self.status_label)
        layout.addLayout(status_layout)
        
        main_widget.setLayout(layout)
    
    def load_index_universe(self):
        """Load index universe from Excel"""
        try:
            drive_path = r"G:\My Drive\PampleMousseLabs\Project Canneberge"
            wb = openpyxl.load_workbook(f"{drive_path}/Beta and Vol module.xlsm")
            ws = wb['Inputs']
            
            # Extract index list (assuming column Q in the screenshot)
            indices = []
            for row in ws.iter_rows(min_row=4, max_row=20, min_col=17, max_col=17):
                val = row[0].value
                if val and isinstance(val, str):
                    indices.append(val.strip())
            
            if not indices:
                indices = ["^GSPC"]  # Fallback
            
            self.index_combo.addItems(indices)
            self.index_combo.setCurrentText("^GSPC")
        except Exception as e:
            self.status_label.setText(f"Error loading indices: {e}")
            self.index_combo.addItem("^GSPC")
    
    def load_tickers_from_excel(self):
        """Load tickers from Project_Canneberge tblIngest"""
        try:
            drive_path = r"G:\My Drive\PampleMousseLabs\Project Canneberge"
            wb = openpyxl.load_workbook(f"{drive_path}/Project_Canneberge.xlsm")
            ws = wb['Control']
            
            tbl_ingest = ws.tables['tblIngest']
            tbl_ref = tbl_ingest.ref
            
            tickers = []
            for row in ws[tbl_ref]:
                val = row[0].value
                if val and isinstance(val, str) and val.upper() != 'TICKER':
                    tickers.append(val.strip())
            
            if tickers:
                self.tickers_input.setText(", ".join(tickers))
        except Exception as e:
            self.status_label.setText(f"Error loading tickers: {e}")
    
    def on_calculate_clicked(self):
        """Handle Calculate button click"""
        if self.worker and self.worker.isRunning():
            self.status_label.setText("Calculation already in progress...")
            return
        
        # Parse inputs
        valuation_date = self.date_input.date().toPyDate()
        beta_history = self.beta_history_input.value()
        vol_term = self.vol_term_input.value()
        index_ticker = self.index_combo.currentText()
        
        tickers_str = self.tickers_input.text()
        tickers = [t.strip() for t in tickers_str.split(",") if t.strip()]
        
        if not tickers:
            self.status_label.setText("Please enter at least one ticker")
            return
        
        self.status_label.setText("Fetching data from Yahoo Finance...")
        self.calculate_btn.setEnabled(False)
        
        # Check if repull is needed
        needs_repull = self.cache.needs_repull(tickers, index_ticker, beta_history, vol_term)
        
        if not needs_repull and self.cache.ticker_prices:
            # Recalc from cache
            self.status_label.setText("Recalculating from cached data...")
            results = self._calculate_from_cache(tickers)
            self._display_results(results)
            self.calculate_btn.setEnabled(True)
        else:
            # Start worker thread
            self.worker = DataPullWorker(tickers, index_ticker, valuation_date, beta_history, vol_term, self.cache)
            self.worker.results.connect(self._display_results)
            self.worker.error.connect(self._on_error)
            self.worker.finished.connect(self._on_finished)
            self.worker.start()
    
    def _calculate_from_cache(self, tickers):
        """Recalculate from cached data"""
        # Reuse worker's calculation methods but from cache
        pass  # TODO: implement cache recalc
    
    def _display_results(self, results):
        """Display results in table"""
        self.results_table.setRowCount(len(results))
        
        for row_idx, result in enumerate(results):
            self.results_table.setItem(row_idx, 0, QTableWidgetItem(result['Ticker']))
            self.results_table.setItem(row_idx, 1, QTableWidgetItem(f"{result['2yr Raw']:.4f}" if result['2yr Raw'] else "N/A"))
            self.results_table.setItem(row_idx, 2, QTableWidgetItem(f"{result['2yr Adj']:.4f}" if result['2yr Adj'] else "N/A"))
            self.results_table.setItem(row_idx, 3, QTableWidgetItem(f"{result['5yr Raw']:.4f}" if result['5yr Raw'] else "N/A"))
            self.results_table.setItem(row_idx, 4, QTableWidgetItem(f"{result['5yr Adj']:.4f}" if result['5yr Adj'] else "N/A"))
            self.results_table.setItem(row_idx, 5, QTableWidgetItem(f"{result['Volatility']:.4f}" if result['Volatility'] else "N/A"))
            self.results_table.setItem(row_idx, 6, QTableWidgetItem(f"{result['YearsAvail']:.2f}"))
            self.results_table.setItem(row_idx, 7, QTableWidgetItem("OK"))
        
        self.status_label.setText(f"Calculated {len(results)} tickers. Last pull: {self.cache.pull_timestamp.strftime('%m/%d/%Y %H:%M:%S')}")
    
    def _on_error(self, error_msg):
        """Handle error from worker thread"""
        self.status_label.setText(f"Error: {error_msg}")
        self.calculate_btn.setEnabled(True)
    
    def _on_finished(self):
        """Handle worker thread completion"""
        self.calculate_btn.setEnabled(True)

# =========================================================
# MAIN
# =========================================================
if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = BetaVolCalculator()
    window.show()
    sys.exit(app.exec())