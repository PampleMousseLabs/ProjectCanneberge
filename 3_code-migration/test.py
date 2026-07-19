import openpyxl

# Path to files on Google Drive
drive_path = r"G:\My Drive\PampleMousseLabs\Project Canneberge"

# Open the Beta/Vol workbook
wb = openpyxl.load_workbook(f"{drive_path}/Beta and Vol module.xlsm")

# Get the sheet names
print("=== SHEETS ===")
for sheet_name in wb.sheetnames:
    print(f"  {sheet_name}")

# Read the Inputs sheet
print("\n=== INPUTS SHEET - KEY VALUES ===")
ws_inputs = wb['Inputs']

inputs_to_read = [
    ('B2', 'Index Name'),
    ('B3', 'Index Ticker'),
    ('B4', 'Valuation Date'),
]

for cell_ref, label in inputs_to_read:
    value = ws_inputs[cell_ref].value
    print(f"  {label}: {value}")

# Read Beta_Vol_Results sheet
print("\n=== BETA_VOL_RESULTS - FIRST 5 ROWS ===")
ws_results = wb['Beta_Vol_Results']

for i, row in enumerate(ws_results.iter_rows(min_row=1, max_row=6, values_only=True), 1):
    if i == 1:
        print(f"  Headers: {row}")
    else:
        print(f"  Row {i-1}: {row}")